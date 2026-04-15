"""
Generate TruStage SmartIngest Data Submission Spec — Excel workbook.
Output: artifacts/truestage-databridge/public/TruStage_SmartIngest_DataSpec.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY       = "1E2D40"
DARK_BLUE  = "2F5496"
MID_BLUE   = "4472C4"
LIGHT_BLUE = "D9E1F2"
ORANGE     = "FF5A1F"   # EXL brand
WHITE      = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MID_GRAY   = "D9D9D9"
RED_BG     = "FFE0DC"
RED_FG     = "C00000"
GREEN_BG   = "E2EFDA"
GREEN_FG   = "375623"
AMBER_BG   = "FFF2CC"
AMBER_FG   = "7F6000"

THIN = Side(border_style="thin", color=MID_GRAY)
MED  = Side(border_style="medium", color="AAAAAA")
ALL_THIN  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOT_MED   = Border(bottom=Side(border_style="medium", color=DARK_BLUE))

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="top", wrap_text=wrap)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def style_header_row(ws, row, bg=DARK_BLUE, fg=WHITE, height=22):
    ws.row_dimensions[row].height = height
    for cell in ws[row]:
        if cell.value is not None or True:
            cell.font  = hdr_font(color=fg)
            cell.fill  = fill(bg)
            cell.alignment = center(wrap=True)
            cell.border = Border(
                left=Side(border_style="thin", color="FFFFFF"),
                right=Side(border_style="thin", color="FFFFFF"),
                bottom=Side(border_style="medium", color=WHITE),
            )

def alternate_rows(ws, start_row, end_row, col_count, bg1=WHITE, bg2=LIGHT_GRAY):
    for r in range(start_row, end_row + 1):
        bg = bg1 if (r - start_row) % 2 == 0 else bg2
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", WHITE, bg1, bg2):
                cell.fill = fill(bg)
            cell.border = ALL_THIN
            if not cell.alignment or not cell.alignment.wrap_text:
                cell.alignment = left()

# ── Data ───────────────────────────────────────────────────────────────────────

CANONICAL_FIELDS = [
    # (field_key, label, category, data_type, required, description, format_notes, example)
    ("memberId",              "Member ID",                  "Member",    "Text",    True,  "Unique member identifier at the credit union",              "Any alphanumeric code up to 50 chars",         "MWF-10001"),
    ("firstName",             "First Name",                 "Member",    "Text",    True,  "Member first name",                                         "Free text, max 100 chars",                     "Alice"),
    ("lastName",              "Last Name",                  "Member",    "Text",    True,  "Member last name",                                          "Free text, max 100 chars",                     "Johnson"),
    ("dateOfBirth",           "Date of Birth",              "Member",    "Date",    True,  "Member date of birth",                                      "YYYY-MM-DD (ISO 8601)",                        "1985-03-22"),
    ("socialSecurityNumber",  "Social Security Number",     "Member",    "Text",    False, "Member SSN — last 4 digits or full 9-digit number",          "9 digits or last-4 format: XXXX",              "6789"),
    ("membershipDate",        "Membership Date",            "Member",    "Date",    False, "Date member joined the credit union",                       "YYYY-MM-DD",                                   "2015-06-01"),
    ("memberStatus",          "Member Status",              "Member",    "Text",    False, "Current membership status",                                 "Active | Inactive | Deceased",                 "Active"),
    ("emailAddress",          "Email Address",              "Contact",   "Email",   False, "Primary email for member communication",                    "Standard email format user@domain.com",        "alice@example.com"),
    ("phoneNumber",           "Phone Number",               "Contact",   "Phone",   False, "Primary phone number",                                      "10 digits, no dashes or spaces: 6085551234",   "6085551234"),
    ("streetAddress",         "Street Address",             "Contact",   "Text",    False, "Street address line 1",                                     "Free text",                                    "123 Oak St"),
    ("city",                  "City",                       "Contact",   "Text",    False, "City of residence",                                         "Free text",                                    "Madison"),
    ("state",                 "State",                      "Contact",   "Text",    False, "US state abbreviation",                                     "2-letter abbreviation: WI, CA, TX",            "WI"),
    ("zipCode",               "ZIP Code",                   "Contact",   "Text",    False, "US postal ZIP code",                                        "5 digits: 53703",                              "53703"),
    ("accountNumber",         "Account Number",             "Account",   "Text",    True,  "Credit union account number",                               "Alphanumeric, up to 50 chars",                 "ACC-00234"),
    ("accountType",           "Account Type",               "Account",   "Text",    False, "Type of account",                                           "Savings | Checking | IRA | 401k | CD | Money Market", "Savings"),
    ("accountOpenDate",       "Account Open Date",          "Account",   "Date",    False, "Date account was opened",                                   "YYYY-MM-DD",                                   "2018-09-15"),
    ("accountBalance",        "Account Balance",            "Financial", "Number",  False, "Current account balance in USD",                            "Decimal number, no $ sign: 5200.00",           "5200.00"),
    ("investmentBalance",     "Investment Balance",         "Financial", "Number",  False, "Total investment portfolio balance",                        "Decimal number: 12500.75",                     "12500.75"),
    ("loanBalance",           "Loan Balance",               "Financial", "Number",  False, "Outstanding loan balance",                                  "Decimal number; omit or use 0 if no loan",     "8400.00"),
    ("retirementBalance",     "Retirement Account Balance", "Financial", "Number",  False, "Total retirement account balance (IRA/401k)",               "Decimal number",                               "47200.00"),
    ("monthlyContribution",   "Monthly Contribution",       "Financial", "Number",  False, "Monthly investment or retirement contribution amount",       "Decimal number between 0 and 100000",          "250.00"),
    ("riskProfile",           "Risk Profile",               "Financial", "Text",    False, "Investor risk tolerance",                                   "Conservative | Moderate | Aggressive",         "Moderate"),
    ("annualIncome",          "Annual Income",              "Financial", "Number",  False, "Reported annual household income",                          "Decimal number",                               "72000.00"),
]

DQ_RULES = [
    # (name, field, rule_type, severity, description, what_happens)
    ("Member ID Required",           "memberId",          "Required",  "Hard", "Every record must have a valid Member ID.",                                   "Records without a Member ID are rejected and reported to the credit union."),
    ("First Name Required",          "firstName",         "Required",  "Hard", "Member first name must be present.",                                          "Records without a First Name are rejected."),
    ("Last Name Required",           "lastName",          "Required",  "Hard", "Member last name must be present.",                                           "Records without a Last Name are rejected."),
    ("Account Number Required",      "accountNumber",     "Required",  "Hard", "Every record must have an account number.",                                   "Records without an Account Number are rejected."),
    ("Valid Email Format",           "emailAddress",      "Format",    "Soft", "Email must follow standard format: user@domain.com",                         "Record is flagged with a soft error but still accepted for processing."),
    ("Valid US Phone",               "phoneNumber",       "Format",    "Soft", "Phone number must be exactly 10 digits with no dashes or spaces.",            "Record is flagged with a soft error."),
    ("Valid ZIP Code",               "zipCode",           "Format",    "Soft", "ZIP code must be exactly 5 numeric digits.",                                  "Record is flagged with a soft error."),
    ("Account Balance Non-Negative", "accountBalance",    "Range",     "Soft", "Account balance must be ≥ 0.",                                               "Record is flagged; negative balances are highlighted in the error report."),
    ("Investment Balance Non-Neg.",  "investmentBalance", "Range",     "Soft", "Investment balance must be ≥ 0.",                                            "Record is flagged with a soft error."),
    ("Monthly Contribution Range",   "monthlyContribution","Range",    "Soft", "Monthly contribution must be between $0 and $100,000.",                      "Record is flagged with a soft error."),
]

SAMPLE_DATA = [
    ("memberId",            "MWF-10001"),
    ("firstName",           "Alice"),
    ("lastName",            "Johnson"),
    ("dateOfBirth",         "1985-03-22"),
    ("socialSecurityNumber","6789"),
    ("membershipDate",      "2015-06-01"),
    ("memberStatus",        "Active"),
    ("emailAddress",        "alice.johnson@email.com"),
    ("phoneNumber",         "6085551234"),
    ("streetAddress",       "123 Oak Street"),
    ("city",                "Madison"),
    ("state",               "WI"),
    ("zipCode",             "53703"),
    ("accountNumber",       "ACC-00234"),
    ("accountType",         "Savings"),
    ("accountOpenDate",     "2018-09-15"),
    ("accountBalance",      "5200.00"),
    ("investmentBalance",   "12500.75"),
    ("loanBalance",         "0.00"),
    ("retirementBalance",   "47200.00"),
    ("monthlyContribution", "250.00"),
    ("riskProfile",         "Moderate"),
    ("annualIncome",        "72000.00"),
]

# ── Build workbook ─────────────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Cover
# ══════════════════════════════════════════════════════════════════════════════
cover = wb.create_sheet("Cover")
cover.sheet_view.showGridLines = False
cover.column_dimensions["A"].width = 4
cover.column_dimensions["B"].width = 60
cover.column_dimensions["C"].width = 24

# Header banner (rows 1-6)
for r in range(1, 7):
    cover.row_dimensions[r].height = 18
    for c in range(1, 4):
        cover.cell(r, c).fill = fill(NAVY)

cover.merge_cells("B2:C2")
t = cover.cell(2, 2, "SmartIngest by EXL")
t.font = Font(name="Calibri", bold=True, size=22, color=ORANGE)
t.alignment = left(wrap=False)

cover.merge_cells("B3:C3")
t = cover.cell(3, 2, "TruStage Wealth Management")
t.font = Font(name="Calibri", bold=False, size=14, color=WHITE)
t.alignment = left(wrap=False)

cover.merge_cells("B4:C4")
t = cover.cell(4, 2, "Credit Union Data Submission Requirements")
t.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
t.alignment = left(wrap=False)

cover.row_dimensions[6].height = 6  # spacer

# Intro section
def cover_row(ws, row, label, value, height=18):
    ws.row_dimensions[row].height = height
    lbl = ws.cell(row, 2, label)
    lbl.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
    lbl.alignment = left(wrap=False)
    val = ws.cell(row, 3, value)
    val.font = Font(name="Calibri", size=10, color="333333")
    val.alignment = left(wrap=False)

cover_row(cover,  8, "Purpose:",        "This workbook specifies the exact format TruStage requires for all member data submissions.")
cover_row(cover,  9, "Version:",        "2025-Q2")
cover_row(cover, 10, "File formats:",   "CSV (.csv) · JSON (.json) · Tab-separated (.tsv)")
cover_row(cover, 11, "Submission via:", "SmartIngest portal — upload directly from the Ingestion Jobs page")
cover_row(cover, 12, "Questions:",      "Contact your TruStage Integration Manager")

# Tabs overview
cover.row_dimensions[14].height = 20
t = cover.cell(14, 2, "This Workbook Contains:")
t.font = Font(name="Calibri", bold=True, size=12, color=NAVY)
t.alignment = left(wrap=False)

tabs_info = [
    ("Field Requirements",  "All 23 canonical fields — name, type, required/optional, format rules, and examples."),
    ("Data Quality Rules",  "Validation rules applied by SmartIngest — hard errors reject records, soft errors flag them."),
    ("Sample Data",         "A correctly formatted example row showing every field populated."),
    ("FAQ",                 "Answers to common questions about formatting, error handling, and re-submissions."),
]
for i, (tab, desc) in enumerate(tabs_info, start=16):
    cover.row_dimensions[i].height = 18
    c = cover.cell(i, 2)
    c.value = f"  ➤  {tab}"
    c.font = Font(name="Calibri", bold=True, size=10, color=MID_BLUE)
    c.alignment = left(wrap=False)
    d = cover.cell(i, 3)
    d.value = desc
    d.font = Font(name="Calibri", size=10, color="444444")
    d.alignment = left(wrap=True)

# Footer
cover.row_dimensions[22].height = 16
f = cover.cell(22, 2, "CONFIDENTIAL — FOR AUTHORIZED CREDIT UNION PARTNERS ONLY")
f.font = Font(name="Calibri", italic=True, size=8, color="999999")


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Field Requirements
# ══════════════════════════════════════════════════════════════════════════════
ws_fields = wb.create_sheet("Field Requirements")
ws_fields.freeze_panes = "A3"
ws_fields.sheet_view.showGridLines = False

FIELD_HEADERS = ["#", "Field Key (column header)", "Display Label", "Category", "Data Type", "Required?", "Description", "Format / Allowed Values", "Example Value"]
FIELD_WIDTHS  = [5,   28,                           22,             13,          12,           11,           42,            38,                         22]

# Title row
ws_fields.merge_cells(f"A1:{get_column_letter(len(FIELD_HEADERS))}1")
title = ws_fields.cell(1, 1, "TruStage SmartIngest — Required Field Specification")
title.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
title.fill = fill(NAVY)
title.alignment = center()
ws_fields.row_dimensions[1].height = 26

# Header row
for j, h in enumerate(FIELD_HEADERS, start=1):
    ws_fields.cell(2, j, h)
for j, w in enumerate(FIELD_WIDTHS, start=1):
    set_col_width(ws_fields, get_column_letter(j), w)
style_header_row(ws_fields, 2, bg=DARK_BLUE, fg=WHITE, height=24)

# Category color map
CAT_COLORS = {"Member": "D9E1F2", "Contact": "E2EFDA", "Account": "FFF2CC", "Financial": "FCE4D6"}

current_cat = None
row = 3
for i, (key, label, cat, dtype, req, desc, fmt, ex) in enumerate(CANONICAL_FIELDS, start=1):
    # Category separator row
    if cat != current_cat:
        current_cat = cat
        ws_fields.row_dimensions[row].height = 16
        ws_fields.merge_cells(f"A{row}:{get_column_letter(len(FIELD_HEADERS))}{row}")
        sep = ws_fields.cell(row, 1, f"  {cat.upper()} FIELDS")
        sep.font = Font(name="Calibri", bold=True, size=9, color=NAVY)
        sep.fill = fill(CAT_COLORS.get(cat, LIGHT_BLUE))
        sep.alignment = Alignment(horizontal="left", vertical="center")
        sep.border = Border(bottom=Side(border_style="thin", color=DARK_BLUE))
        row += 1

    ws_fields.row_dimensions[row].height = 40

    vals = [i, key, label, cat, dtype, "Required" if req else "Optional", desc, fmt, ex]
    for j, v in enumerate(vals, start=1):
        c = ws_fields.cell(row, j, v)
        c.font = body_font(size=10, bold=(j in (2, 3)))
        c.border = ALL_THIN
        c.alignment = left()

        # Color required/optional
        if j == 6:
            if req:
                c.fill = fill(RED_BG)
                c.font = Font(name="Calibri", bold=True, size=10, color=RED_FG)
            else:
                c.fill = fill(GREEN_BG)
                c.font = Font(name="Calibri", size=10, color=GREEN_FG)
            c.alignment = center()
        elif j == 1:
            c.alignment = center()
            c.fill = fill(LIGHT_BLUE)
        else:
            bg = CAT_COLORS.get(cat, WHITE)
            c.fill = fill(bg if j == 4 else WHITE)
    row += 1

# Add Excel table
last_col = get_column_letter(len(FIELD_HEADERS))
tab = Table(displayName="FieldSpec", ref=f"A2:{last_col}{row - 1}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
ws_fields.add_table(tab)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Data Quality Rules
# ══════════════════════════════════════════════════════════════════════════════
ws_dq = wb.create_sheet("Data Quality Rules")
ws_dq.freeze_panes = "A3"
ws_dq.sheet_view.showGridLines = False

DQ_HEADERS = ["#", "Rule Name", "Field", "Rule Type", "Severity", "Description", "What Happens on Failure?"]
DQ_WIDTHS  = [5,   32,          22,       13,           11,          52,             50]

ws_dq.merge_cells(f"A1:{get_column_letter(len(DQ_HEADERS))}1")
t = ws_dq.cell(1, 1, "TruStage SmartIngest — Data Quality Rules")
t.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
t.fill = fill(NAVY)
t.alignment = center()
ws_dq.row_dimensions[1].height = 26

for j, h in enumerate(DQ_HEADERS, start=1):
    ws_dq.cell(2, j, h)
for j, w in enumerate(DQ_WIDTHS, start=1):
    set_col_width(ws_dq, get_column_letter(j), w)
style_header_row(ws_dq, 2, bg=DARK_BLUE, fg=WHITE, height=24)

SREV_COLORS = {"Hard": (RED_BG, RED_FG), "Soft": (AMBER_BG, AMBER_FG)}
TYPE_COLORS  = {"Required": LIGHT_BLUE, "Format": "E8D5FF", "Range": "D5F0FF"}

for i, (name, field, rtype, sev, desc, outcome) in enumerate(DQ_RULES, start=1):
    row_n = i + 2
    ws_dq.row_dimensions[row_n].height = 44
    vals = [i, name, field, rtype, sev, desc, outcome]
    for j, v in enumerate(vals, start=1):
        c = ws_dq.cell(row_n, j, v)
        c.font = body_font(size=10)
        c.border = ALL_THIN
        c.alignment = left()
        if j == 5:
            bg, fg = SREV_COLORS.get(sev, (WHITE, "000000"))
            c.fill = fill(bg)
            c.font = Font(name="Calibri", bold=True, size=10, color=fg)
            c.alignment = center()
        elif j == 4:
            c.fill = fill(TYPE_COLORS.get(rtype, WHITE))
            c.alignment = center()
        elif j == 1:
            c.alignment = center()
            c.fill = fill(LIGHT_BLUE)

# Severity legend
legend_row = len(DQ_RULES) + 4
ws_dq.cell(legend_row, 2, "Severity Guide:").font = Font(name="Calibri", bold=True, size=10, color=NAVY)
ws_dq.cell(legend_row+1, 2, "Hard Error").font = Font(name="Calibri", bold=True, size=10, color=RED_FG)
ws_dq.cell(legend_row+1, 2).fill = fill(RED_BG)
ws_dq.cell(legend_row+1, 3, "Record is rejected and will NOT be loaded into TruStage systems. Included in the error report sent to the credit union.").font = body_font(size=9)
ws_dq.cell(legend_row+2, 2, "Soft Error").font = Font(name="Calibri", bold=True, size=10, color=AMBER_FG)
ws_dq.cell(legend_row+2, 2).fill = fill(AMBER_BG)
ws_dq.cell(legend_row+2, 3, "Record is accepted but flagged for review. Noted in the summary report. No re-submission needed unless requested.").font = body_font(size=9)
for r in range(legend_row, legend_row + 3):
    ws_dq.row_dimensions[r].height = 20


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Sample Data
# ══════════════════════════════════════════════════════════════════════════════
ws_sample = wb.create_sheet("Sample Data")
ws_sample.sheet_view.showGridLines = False

ws_sample.merge_cells("A1:C1")
t = ws_sample.cell(1, 1, "TruStage SmartIngest — Example Correctly Formatted Record")
t.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
t.fill = fill(NAVY)
t.alignment = center()
ws_sample.row_dimensions[1].height = 26

# Headers
for j, hdr in enumerate(["Field Key", "Display Label", "Sample Value", "Notes"], start=1):
    c = ws_sample.cell(2, j, hdr)
    c.font = hdr_font()
    c.fill = fill(DARK_BLUE)
    c.alignment = center()
    c.border = ALL_THIN

set_col_width(ws_sample, "A", 28)
set_col_width(ws_sample, "B", 28)
set_col_width(ws_sample, "C", 22)
set_col_width(ws_sample, "D", 44)
ws_sample.freeze_panes = "A3"

FIELD_LOOKUP = {f[0]: f for f in CANONICAL_FIELDS}
for i, (key, val) in enumerate(SAMPLE_DATA, start=3):
    meta = FIELD_LOOKUP.get(key)
    ws_sample.row_dimensions[i].height = 20
    bg = LIGHT_GRAY if (i % 2 == 0) else WHITE
    for j, v in enumerate([key, meta[1] if meta else key, val, meta[7] if meta else ""], start=1):
        c = ws_sample.cell(i, j, v)
        c.font = body_font(size=10, bold=(j == 3))
        c.fill = fill(bg)
        c.border = ALL_THIN
        c.alignment = left(wrap=False)
    # highlight required fields
    if meta and meta[4]:
        ws_sample.cell(i, 1).font = Font(name="Calibri", bold=True, size=10, color=RED_FG)

ws_sample.row_dimensions[2].height = 22

# How to use CSV note
note_row = len(SAMPLE_DATA) + 4
ws_sample.merge_cells(f"A{note_row}:D{note_row}")
n = ws_sample.cell(note_row, 1, "  HOW TO USE THIS SAMPLE")
n.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
n.fill = fill(LIGHT_BLUE)
n.alignment = left(wrap=False)

notes = [
    "Your file's first row should contain the column header names (Field Key column above) — SmartIngest uses AI to map them automatically.",
    "Column order does NOT matter; SmartIngest will detect and map each field based on its name and sample values.",
    "Columns not listed in this spec will be ignored by SmartIngest. Extra columns are safe to include.",
    "Dates must be formatted as YYYY-MM-DD. Other formats (MM/DD/YYYY etc.) may be detected but YYYY-MM-DD is recommended.",
    "Numeric fields should NOT include $ signs, commas, or currency symbols — plain decimal numbers only (e.g. 5200.00).",
]
for k, note in enumerate(notes, start=1):
    r = note_row + k
    ws_sample.row_dimensions[r].height = 18
    ws_sample.merge_cells(f"A{r}:D{r}")
    c = ws_sample.cell(r, 1, f"  {k}.  {note}")
    c.font = body_font(size=9, color="333333")
    c.alignment = left(wrap=False)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5 — FAQ
# ══════════════════════════════════════════════════════════════════════════════
ws_faq = wb.create_sheet("FAQ")
ws_faq.sheet_view.showGridLines = False
ws_faq.column_dimensions["A"].width = 4
ws_faq.column_dimensions["B"].width = 44
ws_faq.column_dimensions["C"].width = 62

ws_faq.merge_cells("A1:C1")
t = ws_faq.cell(1, 1, "TruStage SmartIngest — Frequently Asked Questions")
t.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
t.fill = fill(NAVY)
t.alignment = center()
ws_faq.row_dimensions[1].height = 26

faqs = [
    ("What file formats are accepted?",
     "CSV (.csv), JSON (.json), and Tab-Separated Values (.tsv). Each file should represent one batch of member records. The maximum file size is 50 MB per upload."),
    ("What should I use as column headers?",
     "Use the 'Field Key' names from the Field Requirements tab (e.g. memberId, firstName). SmartIngest's AI will attempt to match non-standard headers too, but exact matches result in 100% confidence scores and faster processing."),
    ("Which fields are truly required?",
     "Four fields are required and will cause hard errors if missing: memberId, firstName, lastName, and accountNumber. All other fields are optional but recommended for complete member profiles."),
    ("What is a hard error vs. a soft error?",
     "Hard Error: The record is rejected and excluded from the data load. These will be listed in the error report sent to your data team.\nSoft Error: The record is accepted but flagged for review. These are noted in the summary report for your awareness."),
    ("What happens after I upload a file?",
     "SmartIngest runs a 4-stage pipeline: (1) File is parsed, (2) AI maps your columns to the TruStage canonical schema, (3) Data quality rules are applied to every record, (4) If errors exist, an AI-drafted error report email is generated for your team to review and send."),
    ("Can I include extra columns not in this spec?",
     "Yes. Any columns not recognized by the canonical schema will be ignored safely. They do not cause errors."),
    ("How do I fix a hard error rejection?",
     "Correct the records flagged in the error report and re-upload the corrected file as a new submission. Only include the corrected records in the re-submission to avoid duplicates."),
    ("What date format should I use?",
     "ISO 8601 format: YYYY-MM-DD (e.g. 1985-03-22). SmartIngest will attempt to parse other formats but YYYY-MM-DD is strongly recommended to avoid mismatch."),
    ("Should phone numbers include dashes or spaces?",
     "No. Phone numbers should be 10 consecutive digits with no formatting characters. Example: 6085551234. The validation rule will flag records with dashes, spaces, or parentheses."),
    ("What if a member has no email or phone?",
     "Email and phone are optional fields. Leave the column blank or omit it entirely. Do not enter placeholder values like 'N/A' or '0000000000' — these will fail format validation."),
    ("How are balances handled for members with no loan?",
     "Submit 0.00 for loanBalance if no loan exists. Leave the field blank only if the member has no account of that type whatsoever."),
    ("Is my data secure during transmission?",
     "Yes. All uploads are transmitted over HTTPS (TLS 1.2+) and stored in TruStage's secure cloud infrastructure. Data is never shared outside the TruStage platform."),
]

row = 3
for i, (q, a) in enumerate(faqs, start=1):
    # Q row
    ws_faq.row_dimensions[row].height = 20
    q_cell = ws_faq.cell(row, 2, f"Q{i}:  {q}")
    q_cell.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
    q_cell.fill = fill(LIGHT_BLUE)
    q_cell.alignment = left(wrap=False)
    q_cell.border = Border(left=Side(border_style="medium", color=DARK_BLUE), top=Side(border_style="thin", color=MID_GRAY))
    ws_faq.merge_cells(f"B{row}:C{row}")
    row += 1
    # A row
    lines = a.split("\n")
    ws_faq.row_dimensions[row].height = max(18, len(lines) * 17)
    a_cell = ws_faq.cell(row, 2, a)
    a_cell.font = body_font(size=10, color="333333")
    a_cell.alignment = left(wrap=True)
    a_cell.fill = fill(WHITE)
    a_cell.border = Border(
        left=Side(border_style="medium", color=DARK_BLUE),
        bottom=Side(border_style="thin", color=MID_GRAY),
    )
    ws_faq.merge_cells(f"B{row}:C{row}")
    row += 2  # spacer


# ── Save ───────────────────────────────────────────────────────────────────────
os.makedirs("artifacts/truestage-databridge/public", exist_ok=True)
out_path = "artifacts/truestage-databridge/public/TruStage_SmartIngest_DataSpec.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
